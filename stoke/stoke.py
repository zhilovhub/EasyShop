import datetime
import json
import os
import shutil
import csv
from string import ascii_letters, digits
from random import sample

from typing import Iterable

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from database.models.models import Database
from database.models.product_model import ProductNotFound, ProductWithoutId


def singleton(class_):
    instances = {}

    def get_instance(*args, **kwargs):
        if class_ not in instances:
            instances[class_] = class_(*args, **kwargs)
        return instances[class_]

    return get_instance


@singleton
class Stoke:  # TODO raise specific exceptions in import methods + optimize (union) pictures logic
    """Модуль склада"""

    def __init__(self, database: Database) -> None:
        self.product_db = database.get_product_db()
        self.files_path = os.environ["FILES_PATH"]

    async def import_json(
            self, bot_id: int, path_to_file: str, replace: bool, path_to_file_with_pictures: str = None
    ) -> None:
        """If ``replace`` is true then first delete all products else just add or update by name"""
        with open(path_to_file, "r", encoding="utf-8") as f:
            products = map(lambda x: ProductWithoutId(bot_id=bot_id, **x), json.load(f))

        if replace:
            await self.product_db.delete_all_products(bot_id)

        await self._import_products(bot_id, products, replace, path_to_file_with_pictures)

    async def export_json(self, bot_id: int, with_pictures: bool = False) -> tuple[str, str | None]:
        """Экспорт товаров в виде json файла"""
        products = await self.product_db.get_all_products(bot_id)
        if with_pictures:
            path_to_images = self._generate_path_for_pictures(bot_id)
        else:
            path_to_images = None

        json_products = []
        for product in products:
            json_products.append({
                "name": product.name,
                "description": product.description,
                "price": product.price,
                "count": product.count
            })
            if with_pictures:
                picture = product.picture
                json_products[-1]["picture"] = picture
                if picture:
                    shutil.copyfile(self.files_path + picture, path_to_images + picture)

        path_to_file = self._generate_path_to_file(bot_id, "json")
        with open(path_to_file, "w", encoding="utf-8") as f:
            json.dump(json_products, f, indent=4, ensure_ascii=False)

        return path_to_file, path_to_images

    async def import_csv(
            self, bot_id: int, path_to_file: str, replace: bool, path_to_file_with_pictures: str = None
    ) -> None:
        """If ``replace`` is true then first delete all products else just add or update by name"""
        with open(path_to_file, "r", encoding="latin-1") as f:
            delimiter = csv.Sniffer().sniff(f.read(1024)).delimiter
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            next(reader)

            products = []
            for row in reader:
                products.append(ProductWithoutId(
                    bot_id=bot_id,
                    name=row[0],
                    description=row[1] if row[1] is not None else "",
                    price=row[2],
                    count=row[3],
                    picture=row[4]
                ))

        await self._import_products(bot_id, products, replace, path_to_file_with_pictures)

    async def export_csv(self, bot_id: int, with_pictures: bool = False) -> tuple[str, str | None]:
        """Экспорт товаров в виде csv файла"""
        products = await self.product_db.get_all_products(bot_id)
        if with_pictures:
            path_to_images = self._generate_path_for_pictures(bot_id)
        else:
            path_to_images = None

        path_to_file = self._generate_path_to_file(bot_id, "csv")
        with open(path_to_file, "w", newline="") as f:
            writer = csv.writer(f, delimiter=",")
            writer.writerow(["Название", "Описание", "Цена", "Кол-во"] + (["Картинка"] if with_pictures else []))

            for product in products:
                picture = product.picture
                writer.writerow(
                    [product.name, product.description, product.price, product.count] +
                    ([picture] if with_pictures else [])
                )
                if with_pictures and picture:
                    shutil.copyfile(self.files_path + picture, path_to_images + picture)

        return path_to_file, path_to_images

    async def import_xlsx(
            self, bot_id: int, path_to_file: str, replace: bool, path_to_file_with_pictures: str = None
    ) -> None:  # TODO come up with picture
        """If ``replace`` is true then first delete all products else just add or update by name"""
        wb = load_workbook(filename=path_to_file)
        ws = wb.active

        # should be name, description, price, count, picture
        products = []
        for row in list(ws.values)[1:]:
            products.append(ProductWithoutId(
                bot_id=bot_id,
                name=row[0],
                description=row[1] if row[1] is not None else "",
                price=row[2],
                count=row[3],
                picture=row[4] if len(row) > 4 else None
            ))

        await self._import_products(bot_id, products, replace, path_to_file_with_pictures)

    async def export_xlsx(self, bot_id: int, with_pictures: bool = False) -> tuple[str, str | None]:
        """Экспорт товаров в виде Excel файла"""
        products = await self.product_db.get_all_products(bot_id)
        if with_pictures:
            path_to_images = self._generate_path_for_pictures(bot_id)
        else:
            path_to_images = None

        wb = Workbook()
        ws = wb.active

        column_names = ["Название", "Описание", "Цена", "Кол-во"] + (["Картинка"] if with_pictures else [])
        column_ind = ['A', 'B', 'C', 'D'] + (["E"] if with_pictures else [])

        ft = Font()
        ft.bold = True

        for ind, name in zip(column_ind, column_names):
            ws[f'{ind}1'] = name
            ws[f'{ind}1'].font = ft

        for ind, product in enumerate(products, start=2):
            ws[f'A{ind}'] = product.name
            ws[f'B{ind}'] = product.description
            ws[f'C{ind}'] = product.price
            ws[f'D{ind}'] = product.count
            if with_pictures:
                picture = product.picture
                ws[f'E{ind}'] = picture
                if with_pictures and picture:
                    shutil.copyfile(self.files_path + picture, path_to_images + picture)

        path_to_file = self._generate_path_to_file(bot_id, "xlsx")
        wb.save(path_to_file)

        return path_to_file, path_to_images

    async def get_product_count(self, product_id: int) -> int:
        """Возвращает количество товара на складе"""
        try:
            return (await self.product_db.get_product(product_id)).count
        except ProductNotFound:
            return 0

    async def update_product_count(self, product_id: int, new_count: int) -> None:
        """Обновляет количетсво товара на складе"""
        product = await self.product_db.get_product(product_id)
        product.count = new_count
        await self.product_db.update_product(product)

    async def _import_products(
            self, bot_id: int, products: Iterable[ProductWithoutId], replace: bool, path_to_file_with_pictures: str
    ) -> None:
        if replace:
            await self.product_db.delete_all_products(bot_id)
        for product in products:
            if path_to_file_with_pictures:
                self._update_product_picture(product, path_to_file_with_pictures)
            else:
                product.picture = None
            await self.product_db.upsert_product(product)

    def _update_product_picture(self, product: ProductWithoutId, path_to_file_with_pictures: str) -> None:
        new_picture_path = self._generate_path_to_picture()

        with open(new_picture_path, "wb") as f_new_picture:
            with open(f"{path_to_file_with_pictures}{product.picture}", "rb") as f_from_json:
                f_new_picture.write(f_from_json.read())

        product.picture = new_picture_path.split("/")[-1]

    def _generate_path_to_picture(self) -> str:
        return self.files_path + ''.join(sample(digits + ascii_letters, 5)) + ".jpg"

    def _generate_path_to_file(self, bot_id: int, format: str) -> str:
        return self.files_path + \
               f"{bot_id}_" + \
               datetime.datetime.utcnow().strftime("%d%m%y_%H%M%S") + f".{format}"

    def _generate_path_for_pictures(self, bot_id: int) -> str:
        path_to_pictures = self.files_path + \
                           f"{bot_id}_" + \
                           datetime.datetime.utcnow().strftime("%d%m%y_%H%M%S")
        try:
            os.mkdir(path_to_pictures)
            path_to_pictures += "/pictures/"
            os.mkdir(path_to_pictures)
        except Exception as _:
            pass

        return path_to_pictures
