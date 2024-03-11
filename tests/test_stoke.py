import csv
import json
import os
import shutil
from datetime import datetime
from typing import Iterable

from openpyxl import load_workbook

import pytest

from database.models.bot_model import BotDao, BotSchemaWithoutId
from database.models.product_model import ProductDao, ProductWithoutId, ProductSchema
from stoke.stoke import Stoke

BOT_ID = 1

product_schema_without_id_1 = ProductWithoutId(
    bot_id=BOT_ID,
    name="Xbox",
    description="",
    price=21000,
    count=23,
    picture="XYvzR.jpg"
)
product_schema_without_id_2 = ProductWithoutId(
    bot_id=BOT_ID,
    name="Xbox Series X",
    description="",
    price=31000,
    count=12,
    picture=None
)
product_schema_without_id_3 = ProductWithoutId(
    bot_id=BOT_ID,
    name="Xbox Series XXX",
    description="",
    price=55000,
    count=122,
    picture=None
)
product_schema_2 = ProductSchema(
    id=4,
    **product_schema_without_id_2.model_dump()
)
product_schema_3 = ProductSchema(
    id=3,
    **product_schema_without_id_3.model_dump()
)

bot_schema_without_id = BotSchemaWithoutId(
    token="1000000000:AaAaAaAaAaAaAaAaAaAaAaAaAaAaAaAaAaA",
    status="a",
    created_at=datetime.utcnow(),
    created_by=1,
    locale=""
)


@pytest.fixture
async def before_add_two_products(bot_db: BotDao, product_db: ProductDao) -> None:
    await bot_db.add_bot(bot_schema_without_id)
    await product_db.add_product(product_schema_without_id_1)
    await product_db.add_product(product_schema_without_id_2)


class TestStoke:  # TODO optimize import tests + create fixture for cleaning and backing data
    async def test_import_json(self, stoke: Stoke, product_db: ProductDao, before_add_two_products) -> None:
        """Stoke.import_json"""
        await stoke.import_json(
            BOT_ID,
            "tests/raw_files/import_json_1_file_test.json",
            replace=True  # with replace True
        )
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 1
        assert products[0] == product_schema_3

        await stoke.import_json(
            BOT_ID,
            "tests/raw_files/import_json_2_files_test.json",
            replace=False  # with replace False
        )
        product_schema_3.count = 50
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 2
        assert products[0] == product_schema_3 and products[1] == product_schema_2

        # back values
        product_schema_3.count = 122

        await stoke.import_json(
            BOT_ID,
            "tests/raw_files/import_json_1_file_with_picture_test.json",
            replace=False,
            path_to_file_with_pictures="tests/raw_files/product_pictures_test/"  # with pictures
        )

        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 2
        product_picture = products[0].picture
        assert products[1].picture is None and product_picture != "XYvzR.jpg" and \
               len(product_picture.split(".")[0]) == 5 and product_picture.split(".")[1] == "jpg"
        assert self._check_pictures_exists(products, os.environ["FILES_PATH"])

    async def test_export_json(self, stoke: Stoke, before_add_two_products) -> None:
        """Stoke.export_json"""
        path_to_file, _ = await stoke.export_json(bot_id=BOT_ID)
        assert open(path_to_file, "r", encoding="utf-8").read() == \
               open("tests/raw_files/export_json_test.json", "r", encoding="utf-8").read()

        os.remove(path_to_file)

        path_to_file, path_to_images = await stoke.export_json(bot_id=BOT_ID, with_pictures=True)   # with pictures
        assert open(path_to_file, "r", encoding="utf-8").read() == \
               open("tests/raw_files/export_json_with_pictures_test.json", "r", encoding="utf-8").read()

        with open(path_to_file, "r", encoding="utf-8") as f:
            assert self._check_pictures_exists(
                map(lambda x: ProductWithoutId(bot_id=BOT_ID, **x), json.load(f)), path_to_images
            )
        os.remove(path_to_file)
        shutil.rmtree("/".join(path_to_images.split("/")[:-2]))

    async def test_import_csv(self, stoke: Stoke, product_db: ProductDao, before_add_two_products) -> None:
        """Stoke.import_csv"""
        await stoke.import_csv(
            BOT_ID,
            "tests/raw_files/import_csv_1_file_test.csv",
            replace=True  # with replace True
        )
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 1
        assert products[0] == product_schema_3

        await stoke.import_csv(
            BOT_ID,
            "tests/raw_files/import_csv_2_files_test.csv",
            replace=False  # with replace False
        )
        product_schema_3.count = 50
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 2
        assert products[0] == product_schema_3 and products[1] == product_schema_2

        # back values
        product_schema_3.count = 122

        await stoke.import_csv(
            BOT_ID,
            "tests/raw_files/import_csv_1_file_with_picture_test.csv",
            replace=False,
            path_to_file_with_pictures="tests/raw_files/product_pictures_test/"  # with pictures
        )
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 2
        product_picture = products[0].picture
        assert products[1].picture is None and product_picture != "XYvzR.jpg" and \
               len(product_picture.split(".")[0]) == 5 and product_picture.split(".")[1] == "jpg"
        assert self._check_pictures_exists(products, os.environ["FILES_PATH"])

    async def test_export_csv(self, stoke: Stoke, before_add_two_products) -> None:
        """Stoke.export_csv"""
        path_to_file, _ = await stoke.export_csv(BOT_ID)

        with open(path_to_file, "r") as f:
            delimiter = csv.Sniffer().sniff(f.read(1024)).delimiter
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            next(reader)

            product_schema_without_id_1.picture = None
            excepted_products = [product_schema_without_id_1, product_schema_without_id_2]
            for excepted_product, row in zip(excepted_products, reader):
                assert ProductWithoutId(
                    bot_id=BOT_ID,
                    name=row[0],
                    description=row[1] if row[1] is not None else "",
                    price=row[2],
                    count=row[3],
                    picture=None
                ) == excepted_product

        os.remove(path_to_file)

        # back values
        product_schema_without_id_1.picture = "XYvzR.jpg"

        path_to_file, path_to_images = await stoke.export_csv(BOT_ID, with_pictures=True)
        with open(path_to_file, "r") as f:
            delimiter = csv.Sniffer().sniff(f.read(1024)).delimiter
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            next(reader)

            excepted_products = [product_schema_without_id_1, product_schema_without_id_2]
            for excepted_product, row in zip(excepted_products, reader):
                assert ProductWithoutId(
                    bot_id=BOT_ID,
                    name=row[0],
                    description=row[1] if row[1] is not None else "",
                    price=row[2],
                    count=row[3],
                    picture=row[4] if row[4] else None
                ) == excepted_product

            assert self._check_pictures_exists(excepted_products, path_to_images)

        os.remove(path_to_file)
        shutil.rmtree("/".join(path_to_images.split("/")[:-2]))

    async def test_import_xlsx(self, stoke: Stoke, product_db: ProductDao, before_add_two_products) -> None:
        """Stoke.import_xlsx"""
        await stoke.import_xlsx(
            BOT_ID,
            "tests/raw_files/import_xlsx_1_file_test.xlsx",
            replace=True  # with replace True
        )
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 1
        assert products[0] == product_schema_3

        await stoke.import_xlsx(
            BOT_ID,
            "tests/raw_files/import_xlsx_2_files_test.xlsx",
            replace=False  # with replace False
        )
        product_schema_3.count = 50
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 2
        assert products[0] == product_schema_3 and products[1] == product_schema_2

        # back values
        product_schema_3.count = 122

        await stoke.import_xlsx(
            BOT_ID,
            "tests/raw_files/import_xlsx_1_file_with_picture_test.xlsx",
            replace=False,
            path_to_file_with_pictures="tests/raw_files/product_pictures_test/"  # with pictures
        )
        products = await product_db.get_all_products(BOT_ID)
        assert len(products) == 2
        product_picture = products[0].picture
        assert products[1].picture is None and product_picture != "XYvzR.jpg" and \
               len(product_picture.split(".")[0]) == 5 and product_picture.split(".")[1] == "jpg"
        assert self._check_pictures_exists(products, os.environ["FILES_PATH"])

    async def test_export_xlsx(self, stoke: Stoke, before_add_two_products) -> None:
        """Stoke.export_xlsx"""
        path_to_file, _ = await stoke.export_xlsx(BOT_ID)
        wb = load_workbook(filename=path_to_file, read_only=True)
        ws = wb.active

        product_schema_without_id_1.picture = None
        excepted_products = [product_schema_without_id_1, product_schema_without_id_2]
        for excepted_product, row in zip(excepted_products, list(ws.values)[1:]):
            assert ProductWithoutId(
                bot_id=BOT_ID,
                name=row[0],
                description=row[1] if row[1] is not None else "",
                price=row[2],
                count=row[3],
                picture=None if len(row) < 5 else row[4]
            ) == excepted_product

        wb.close()
        os.remove(path_to_file)

        # back values
        product_schema_without_id_1.picture = "XYvzR.jpg"

        path_to_file, path_to_images = await stoke.export_xlsx(BOT_ID, with_pictures=True)  # with pictures
        wb = load_workbook(filename=path_to_file, read_only=True)
        ws = wb.active

        excepted_products = [product_schema_without_id_1, product_schema_without_id_2]
        for excepted_product, row in zip(excepted_products, list(ws.values)[1:]):
            assert ProductWithoutId(
                bot_id=BOT_ID,
                name=row[0],
                description=row[1] if row[1] is not None else "",
                price=row[2],
                count=row[3],
                picture=row[4]
            ) == excepted_product

        assert self._check_pictures_exists(excepted_products, path_to_images)
        wb.close()
        os.remove(path_to_file)
        shutil.rmtree("/".join(path_to_images.split("/")[:-2]))

    async def test_get_product_count(self, stoke: Stoke, before_add_two_products) -> None:
        """Stoke.get_product_count"""
        assert await stoke.get_product_count(1) == 23 and await stoke.get_product_count(-20) == 0

    async def test_update_product_count(self, stoke: Stoke, product_db: ProductDao, before_add_two_products) -> None:
        """Stoke.update_product_count"""
        product_id = 1
        count = 345
        await stoke.update_product_count(product_id, count)
        assert (await product_db.get_product(product_id)).count == count

    def _check_pictures_exists(self, products: Iterable[ProductWithoutId], path_to_file_with_pictures: str) -> bool:
        for product in products:
            if product.picture:
                if not os.path.exists(path_to_file_with_pictures + product.picture):
                    return False
        return True
