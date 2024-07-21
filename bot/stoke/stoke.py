import datetime
import json
import os
import shutil
import csv
from random import sample
from string import ascii_letters, digits

from typing import Iterable

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from common_utils.singleton import singleton
from common_utils.env_config import FILES_PATH

from database.config import category_db, product_db
from database.models.models import Database
from database.models.product_model import ProductNotFound, ProductWithoutId
from database.models.category_model import CategorySchemaWithoutId, SameCategoryNameAlreadyExists


class UnknownFileExstension(Exception):
    def __init__(self, message):
        super.__init__(message)


@singleton
class Stoke:
    """Модуль склада"""

    def __init__(self, database: Database) -> None:
        self.product_db = database.get_product_db()
        self.files_path = FILES_PATH

    async def import_json(
            self,
            bot_id: int,
            path_to_file: str,
            replace: bool,
            path_to_file_with_pictures: str = None,
            replace_duplicates: bool = False
    ) -> None:
        """If ``replace`` is true then first delete all products else just add or update by name"""
        with open(path_to_file, "r", encoding="utf-8") as f:
            products = map(lambda x: ProductWithoutId(bot_id=bot_id, **x), json.load(f))

        if replace:
            await self.product_db.delete_all_products(bot_id)

        await self._import_products(
            bot_id,
            products,
            replace,
            path_to_file_with_pictures,
            replace_duplicates=replace_duplicates
        )

    async def export_json(self, bot_id: int, with_pictures: bool = False) -> tuple[str, str | None]:
        """Экспорт товаров в виде json файла"""
        products = await self.product_db.get_all_products(bot_id)
        if with_pictures:
            path_to_images = self._generate_path_for_pictures(bot_id)
        else:
            path_to_images = None

        json_products = []
        for product in products:
            json_products.append({
                "name": product.name,
                "description": product.description,
                "price": product.price,
                "count": product.count
            })
            if with_pictures:
                picture = product.picture[0]
                json_products[-1]["picture"] = picture
                if picture:
                    shutil.copyfile(self.files_path + picture, path_to_images + picture)

        path_to_file = self._generate_path_to_file(bot_id, "json")
        with open(path_to_file, "w", encoding="utf-8") as f:
            json.dump(json_products, f, indent=4, ensure_ascii=False)

        return path_to_file, path_to_images

    async def import_csv(
            self,
            bot_id: int,
            path_to_file: str,
            replace: bool,
            path_to_file_with_pictures: str = None,
            replace_duplicates: bool = False
    ) -> None:
        """If ``replace`` is true then first delete all products else just add or update by name"""
        with open(path_to_file, "r", encoding="latin-1") as f:
            delimiter = csv.Sniffer().sniff(f.read(1024)).delimiter
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            next(reader)

            products = []
            for row in reader:
                products.append(ProductWithoutId(
                    bot_id=bot_id,
                    name=row[0],
                    description=row[1] if row[1] is not None else "",
                    price=int(row[2]),
                    count=int(row[3]),
                    picture=[row[4]]
                ))

        await self._import_products(
            bot_id,
            products,
            replace,
            path_to_file_with_pictures,
            replace_duplicates=replace_duplicates
        )

    async def export_csv(self, bot_id: int, with_pictures: bool = False) -> tuple[str, str | None]:
        """Экспорт товаров в виде csv файла"""
        products = await self.product_db.get_all_products(bot_id)
        if with_pictures:
            path_to_images = self._generate_path_for_pictures(bot_id)
        else:
            path_to_images = None

        path_to_file = self._generate_path_to_file(bot_id, "csv")
        with open(path_to_file, "w", newline="") as f:
            writer = csv.writer(f, delimiter=",")
            writer.writerow(["Название", "Описание", "Цена", "Кол-во"] + (["Картинка"] if with_pictures else []))

            for product in products:
                picture = product.picture[0]
                writer.writerow(
                    [product.name, product.description, product.price, product.count] +
                    ([picture] if with_pictures else [])
                )
                if with_pictures and picture:
                    shutil.copyfile(self.files_path + picture, path_to_images + picture)

        return path_to_file, path_to_images

    async def update_count_xlsx(self, path_to_file: str):
        """Expects the xlsx file without pictures column"""
        status, err_message = await self.check_xlsx(path_to_file)
        if status is False:
            return False, err_message
        wb = load_workbook(filename=path_to_file)
        ws = wb.active
        for ind, row in enumerate(list(ws.values)[1:]):
            try:
                product = await product_db.get_product_by_article(row[4])
                if product.count != int(row[3]):
                    product.count = int(row[3])
                    await product_db.update_product(product)
            except ProductNotFound:
                continue
        return True, ""

    @staticmethod
    async def check_xlsx(path_to_file: str):
        wb = load_workbook(filename=path_to_file)
        ws = wb.active
        for ind, row in enumerate(list(ws.values)[1:]):
            err_message = f"Строка {ind+1}: "
            if len(row) != 6:
                return False, err_message + "Неверное количество колонок"
            try:
                price = int(row[2])
                count = int(row[3])
                if price < 0:
                    return False, err_message + "цена меньше 0"
                elif count < 0:
                    return False, err_message + "остаток товара меньше 0"
                return True, ""
            except ValueError:
                return False, err_message + "остаток или цена - не числа"

    async def import_xlsx(
            self,
            bot_id: int,
            path_to_file: str,
            replace: bool,
            path_to_file_with_pictures: str = None,
            replace_duplicates: bool = False
    ) -> None:
        """If ``replace`` is true then first delete all products else just add or update by name"""
        wb = load_workbook(filename=path_to_file)
        ws = wb.active

        # should be name, description, price, count, picture, article, category
        products = []
        for row in list(ws.values)[1:]:
            try:
                cat_id = await category_db.add_category(CategorySchemaWithoutId(bot_id=bot_id, name=row[5]))
            except SameCategoryNameAlreadyExists as err:
                cat_id = err.cat_id
            products.append(ProductWithoutId(
                article=row[4],
                category=[cat_id],
                bot_id=bot_id,
                name=row[0],
                description=row[1] if row[1] is not None else "",
                price=int(row[2]),
                count=int(row[3])
            ))

        await self._import_products(
            bot_id,
            products,
            replace,
            path_to_file_with_pictures,
            replace_duplicates=replace_duplicates
        )

    async def export_xlsx(self, bot_id: int, with_pictures: bool = False) -> tuple[str, str | None]:
        """Экспорт товаров в виде Excel файла"""
        products = await self.product_db.get_all_products(bot_id)
        if with_pictures:
            path_to_images = self._generate_path_for_pictures(bot_id)
        else:
            path_to_images = None

        wb = Workbook()
        ws = wb.active

        column_names = ["Название", "Описание", "Цена", "Кол-во"] + (["Картинка"] if with_pictures else [])
        column_ind = ['A', 'B', 'C', 'D'] + (["E"] if with_pictures else [])

        ft = Font()
        ft.bold = True

        for ind, name in zip(column_ind, column_names):
            ws[f'{ind}1'] = name
            ws[f'{ind}1'].font = ft

        for ind, product in enumerate(products, start=2):
            ws[f'A{ind}'] = product.name
            ws[f'B{ind}'] = product.description
            ws[f'C{ind}'] = product.price
            ws[f'D{ind}'] = product.count
            if with_pictures:
                picture = product.picture[0]
                ws[f'E{ind}'] = picture
                if with_pictures and picture:
                    shutil.copyfile(self.files_path + picture, path_to_images + picture)

        path_to_file = self._generate_path_to_file(bot_id, "xlsx")
        wb.save(path_to_file)

        return path_to_file, path_to_images

    async def get_product_count(self, product_id: int) -> int:
        """Возвращает количество товара на складе"""
        try:
            return (await self.product_db.get_product(product_id)).count
        except ProductNotFound:
            return 0

    async def update_product_count(self, product_id: int, new_count: int) -> None:
        """Обновляет количетсво товара на складе"""
        product = await self.product_db.get_product(product_id)
        product.count = new_count
        await self.product_db.update_product(product)

    async def _import_products(
            self,
            bot_id: int,
            products: Iterable[ProductWithoutId],
            replace: bool,
            path_to_file_with_pictures: str = None,
            replace_duplicates: bool = False
    ) -> None:
        if replace:
            await self.product_db.delete_all_products(bot_id)
        for product in products:
            if path_to_file_with_pictures:
                self._update_product_picture(product, path_to_file_with_pictures)
            else:
                product.picture = None
            await self.product_db.upsert_product(product, replace_duplicates)

    def _update_product_picture(self, product: ProductWithoutId, path_to_file_with_pictures: str) -> None:
        new_picture_path = self._generate_path_to_picture()

        with open(new_picture_path, "wb") as f_new_picture:
            with open(f"{path_to_file_with_pictures}{product.picture}", "rb") as f_from_json:
                f_new_picture.write(f_from_json.read())

        product.picture = new_picture_path.split("/")[-1]

    def _generate_path_to_picture(self) -> str:
        return self.files_path + ''.join(sample(digits + ascii_letters, 5)) + ".jpg"

    def _generate_path_to_file(self, bot_id: int, file_format: str) -> str:
        return self.files_path + \
            f"{bot_id}_" + \
            datetime.datetime.utcnow().strftime("%d%m%y_%H%M%S") + f".{file_format}"

    def _generate_path_for_pictures(self, bot_id: int) -> str:
        path_to_pictures = self.files_path + \
            f"{bot_id}_" + \
            datetime.datetime.utcnow().strftime("%d%m%y_%H%M%S")
        try:
            os.mkdir(path_to_pictures)
            path_to_pictures += "/pictures/"
            os.mkdir(path_to_pictures)
        except Exception as _:  # noqa
            pass

        return path_to_pictures
