import os
from typing import List
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from io import BytesIO

from aiogram import Bot
from aiogram.types import BufferedInputFile

from common_utils.config import common_settings

from database.config import bot_db, contest_db, category_db, product_db
from database.models.bot_model import BotNotFoundError
from database.models.category_model import CategoryNotFoundError
from database.models.contest_model import ContestUserSchema, ContestNotFoundError
from database.models.product_model import ProductSchema

from bot.main import bot as main_bot

from logs.config import logger, extra_params


def _create_zip_buffer(images) -> BufferedInputFile:
    """
    :return: Zip File with images
    """
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w") as zip_file:
        for file_path in images:
            file_name = os.path.basename(file_path)
            with open(file_path, "rb") as file:
                zip_file.writestr(file_name, file.read())
    zip_buffer.seek(0)
    buffer_file = BufferedInputFile(zip_buffer.read(), filename="images.zip")
    return buffer_file


def _make_xlsx_buffer(name: str, wb_data) -> BufferedInputFile:
    """
    :return: xlsx file with data
    """
    wb = create_excel(wb_data, name, name)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer_file = BufferedInputFile(buffer.read(), filename=f"{name}.xlsx")
    buffer.close()
    return buffer_file


def create_excel(data: dict, sheet_name, table_type) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(data[0].keys())
    ws.append(headers)
    for el in data:
        ws.append([el[header] for header in headers])
    # only for styling
    alignment = Alignment(horizontal="center", vertical="center")
    match table_type:
        case "demo_import":
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    if cell.value:
                        cell.alignment = alignment
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width
        case "banned":
            pass

        case "contest":
            pass

        case "product_export":
            header_fill = PatternFill(fgColor="FFCCFFCC", patternType="solid", fill_type="solid")
            even_fill = PatternFill(fgColor="FFE6FFCC", patternType="solid", fill_type="solid")
            odd_fill = PatternFill(fgColor="FFFFF2CC", patternType="solid", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            )
            bold_font = Font(bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = alignment
                cell.border = thin_border

            # Применяем стили к остальным строкам
            for row in range(2, len(data) + 2):  # Начинаем с 2-й строки до последней строки данных
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    if row % 2 == 0:
                        cell.fill = even_fill
                    else:
                        cell.fill = odd_fill
                    cell.alignment = alignment
                    cell.border = thin_border

            # Устанавливаем ширину колонок
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width
        case _:
            pass

    return wb


async def send_demo_import_xlsx(bot_id):
    try:
        bot = await bot_db.get_bot(bot_id=bot_id)
        created_by = bot.created_by
    except BotNotFoundError as e:
        logger.error(
            f"Provided to excel function bot not found bot_id={bot_id}", extra_params(bot_id=bot_id), exc_info=e
        )
        return
    wb_data = [
        {"Имя": "шаблон", "Описание": "шаблон", "Цена": 0, "Остаток": 0, "Артикул": "шаблон", "Категория": "Шаблон"}
    ]
    buffered_file = _make_xlsx_buffer("demo_import", wb_data)
    await main_bot.send_document(created_by, document=buffered_file, caption="шаблон.xlsx")


async def send_ban_users_xlsx(users_list: List[int], bot_id: int):
    try:
        bot = await bot_db.get_bot(bot_id=bot_id)
        created_by = bot.created_by
    except BotNotFoundError as e:
        logger.error(
            f"Provided to excel function bot not found bot_id={bot_id}", extra_params(bot_id=bot_id), exc_info=e
        )
        return
    custom_bot = Bot(token=bot.token)
    wb_data = []
    for user in users_list:
        chat = await custom_bot.get_chat(user)
        username = chat.username
        wb_data.append(
            {
                "user_id": user,
                "username": username,
            }
        )
    buffered_file = _make_xlsx_buffer("banned", wb_data)
    await main_bot.send_document(created_by, document=buffered_file, caption="ban_users.xlsx")


async def send_contest_results_xlsx(users: list[ContestUserSchema], contest_id: int):
    try:
        contest = await contest_db.get_contest_by_contest_id(contest_id)
    except ContestNotFoundError:
        logger.error(
            f"Provided to excel function contest not found contest_id={contest_id}", extra_params(contest_id=contest_id)
        )
        return

    try:
        bot = await bot_db.get_bot(bot_id=contest.bot_id)
        created_by = bot.created_by
    except BotNotFoundError:
        logger.error(
            f"Provided to excel function bot not found bot_id={contest.bot_id}", extra_params(bot_id=contest.bot_id)
        )
        return

    wb_data = []
    for user in users:
        wb_data.append(
            {
                "user_id": user.user_id,
                "username": "@" + str(user.username),
                "full_name": user.full_name,
                "took part time": user.join_date,
                "won": user.is_won,
            }
        )
    buffered_file = _make_xlsx_buffer("contest", wb_data)
    await main_bot.send_document(created_by, document=buffered_file, caption="Список участников конкурса.")


async def send_products_info_xlsx(bot_id: int, products: list[ProductSchema], with_pics: bool):
    wb_data = []
    images = []

    try:
        bot = await bot_db.get_bot(bot_id=bot_id)
        created_by = bot.created_by
    except BotNotFoundError:
        logger.error(
            f"bot_id={bot_id}: Provided to excel function bot not found bot_id={bot_id}", extra_params(bot_id=bot_id)
        )
        return

    for product in products:
        categories = []
        unexisted_categories = set()
        if product.category and product.category != [0]:
            for cat_id in product.category:
                try:
                    cat_obj = await category_db.get_category(cat_id)
                except CategoryNotFoundError as e:
                    logger.warning(
                        f"bot_id={bot_id}: got unreal category_id={cat_id}",
                        exc_info=e,
                        extra=extra_params(bot_id=bot_id, category_id=cat_id),
                    )
                    if cat_id != 0:
                        unexisted_categories.add(cat_id)
                    continue

                categories.append(cat_obj)
            categories_text = "/".join(map(lambda x: x.name, categories))

            # cleaning unexisted categories
            if unexisted_categories:
                product.category = list(filter(lambda x: x not in unexisted_categories, product.category))
                await product_db.update_product(product)
        else:
            categories_text = "Категория не задана"

        if product.picture:
            images_text = "/".join(product.picture)
            images.extend([common_settings.FILES_PATH + prod for prod in product.picture])
        else:
            images_text = "Картинки не найдены"
        current_line = {
            "Имя": product.name,
            "Описание": product.description,
            "Цена": product.price,
            "Остаток": product.count,
            "Артикул": product.article,
            "Категория": categories_text,
        }
        if with_pics:
            current_line.update({"Картинки товара": images_text})
        wb_data.append(current_line)

    if len(images) != 0 and with_pics:
        buffered_zip_file = _create_zip_buffer(images)
        await main_bot.send_document(created_by, document=buffered_zip_file, caption="Картинки ваших товаров")

    buffered_file = _make_xlsx_buffer("product_export", wb_data)
    await main_bot.send_document(created_by, document=buffered_file, caption="Список товаров загруженных в систему")
